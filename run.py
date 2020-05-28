#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
@author mybsdc <mybsdc@gmail.com>
@date 2020/5/18
@time 12:17
"""

import sys
import time
from openpyxl import load_workbook

target_file_name = r'C:\Users\luolongf\Desktop\缴费统计册.xlsx'

origin_ws = load_workbook(r'C:\Users\luolongf\Desktop\学生名册修复版（所有文字可复制）.xlsx')['Sheet1']
target_wb = load_workbook(target_file_name)
target_ws = target_wb['Sheet1']
all_origin_cells = origin_ws['A3':'K44']
all_target_name_cells = target_ws['B5':'B47']
print('已读取所需表格信息，开始处理...')

name_dict = {}
for name_cell in all_target_name_cells:
    name_dict[name_cell[0].value.strip()] = name_cell[0].row

for row in all_origin_cells:
    name = row[3].value.strip()
    grade = row[1].value.strip()
    is_class = row[2].value
    sex = row[4].value.strip()
    id = row[5].value
    parent_name = row[6].value.strip()
    relation = row[7].value.strip()
    parent_id = row[8].value
    addr = row[9].value.strip()
    tel = row[10].value

    try:
        print('正在写入 {} 的信息到新表格'.format(name))
        row_num = name_dict[name]
        target_ws[f'C{row_num}'] = grade
        target_ws[f'D{row_num}'] = is_class
        target_ws[f'E{row_num}'] = sex
        target_ws[f'F{row_num}'] = id
        target_ws[f'G{row_num}'] = parent_name
        target_ws[f'H{row_num}'] = relation
        target_ws[f'I{row_num}'] = parent_id
        target_ws[f'J{row_num}'] = addr
        target_ws[f'K{row_num}'] = tel

        time.sleep(0.2)
    except Exception as e:
        print(f'有一处错误：{str(e)}')

target_wb.save(target_file_name)
print('恭喜，已处理完成')

pass
